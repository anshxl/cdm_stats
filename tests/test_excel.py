import sqlite3
import io
import os
import pytest
from openpyxl import load_workbook
from cdm_stats.db.schema import create_tables
from cdm_stats.ingestion.seed import seed_teams, seed_maps
from cdm_stats.ingestion.csv_loader import ingest_csv
from cdm_stats.export.excel import export_map_matrix
from cdm_stats.export.excel import export_matchup_prep


MATCH_CSV = """date,team1,team2,two_v_two_winner,slot,map_name,winner,winner_score,loser_score
2026-01-15,DVS,OUG,DVS,1,Tunisia,DVS,6,3
2026-01-15,DVS,OUG,DVS,2,Summit,OUG,250,220
2026-01-15,DVS,OUG,DVS,3,Raid,DVS,3,1
2026-01-15,DVS,OUG,DVS,4,Slums,DVS,6,2"""


@pytest.fixture
def db():
    conn = sqlite3.connect(":memory:")
    create_tables(conn)
    seed_teams(conn)
    seed_maps(conn)
    ingest_csv(conn, io.StringIO(MATCH_CSV))
    yield conn
    conn.close()


def test_export_map_matrix_creates_file(db, tmp_path):
    output_path = tmp_path / "matrix.xlsx"
    export_map_matrix(db, str(output_path))
    assert output_path.exists()


def test_export_map_matrix_has_correct_sheet(db, tmp_path):
    output_path = tmp_path / "matrix.xlsx"
    export_map_matrix(db, str(output_path))
    wb = load_workbook(str(output_path))
    assert "Map Matrix" in wb.sheetnames


def test_export_map_matrix_has_team_rows(db, tmp_path):
    output_path = tmp_path / "matrix.xlsx"
    export_map_matrix(db, str(output_path))
    wb = load_workbook(str(output_path))
    ws = wb["Map Matrix"]
    # Row 1 is header, rows 2-15 are teams
    team_cells = [ws.cell(row=r, column=1).value for r in range(2, 16)]
    assert all(t is not None for t in team_cells)
    assert len(team_cells) == 14


def test_export_matchup_creates_file(db, tmp_path):
    output_path = tmp_path / "matchup_DVS_vs_OUG.xlsx"
    dvs = db.execute("SELECT team_id FROM teams WHERE abbreviation = 'DVS'").fetchone()[0]
    oug = db.execute("SELECT team_id FROM teams WHERE abbreviation = 'OUG'").fetchone()[0]
    export_matchup_prep(db, dvs, oug, str(output_path))
    assert output_path.exists()


def test_export_matchup_has_correct_sheet(db, tmp_path):
    output_path = tmp_path / "matchup_DVS_vs_OUG.xlsx"
    dvs = db.execute("SELECT team_id FROM teams WHERE abbreviation = 'DVS'").fetchone()[0]
    oug = db.execute("SELECT team_id FROM teams WHERE abbreviation = 'OUG'").fetchone()[0]
    export_matchup_prep(db, dvs, oug, str(output_path))
    wb = load_workbook(str(output_path))
    assert "Match-Up Prep" in wb.sheetnames


def test_matchup_prep_includes_ban_section_when_bans_exist(db_with_tournament_match, tmp_path):
    """When tournament/playoff bans exist between two teams, matchup prep includes ban summary."""
    conn = db_with_tournament_match
    from cdm_stats.db.queries import get_team_id_by_abbr

    elv_id = get_team_id_by_abbr(conn, "ELV")
    alu_id = get_team_id_by_abbr(conn, "ALU")

    output_path = tmp_path / "matchup_ban_test.xlsx"
    export_matchup_prep(conn, elv_id, alu_id, str(output_path))
    wb = load_workbook(str(output_path))
    ws = wb.active
    # Find "Ban" text in any cell
    all_values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
    assert any("Ban" in str(v) for v in all_values), "Expected ban summary section in matchup prep"
