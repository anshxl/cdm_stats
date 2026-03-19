import sqlite3
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from cdm_stats.metrics.avoidance import pick_win_loss, defend_win_loss, avoidance_index, target_index
from cdm_stats.metrics.elo import get_current_elo, is_low_confidence
from cdm_stats.metrics.margin import score_margins, dominance_flag
from cdm_stats.db.queries import get_ban_summary, get_team_map_wl, get_team_ban_summary

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
LOW_SAMPLE_THRESHOLD = 4


def _get_all_teams(conn: sqlite3.Connection) -> list[tuple[int, str]]:
    return conn.execute("SELECT team_id, abbreviation FROM teams ORDER BY abbreviation").fetchall()


def _get_all_maps(conn: sqlite3.Connection) -> list[tuple[int, str, str]]:
    return conn.execute(
        "SELECT map_id, map_name, mode FROM maps ORDER BY mode, map_name"
    ).fetchall()


def _cell_color(pick_wl: dict, defend_wl: dict, avoid: dict, tgt: dict) -> PatternFill | None:
    total_sample = (pick_wl["wins"] + pick_wl["losses"] +
                    defend_wl["wins"] + defend_wl["losses"])
    if total_sample == 0:
        return YELLOW_FILL
    if avoid.get("opportunities", 0) < LOW_SAMPLE_THRESHOLD:
        return YELLOW_FILL

    pick_total = pick_wl["wins"] + pick_wl["losses"]
    defend_total = defend_wl["wins"] + defend_wl["losses"]
    pick_rate = pick_wl["wins"] / pick_total if pick_total else 0
    defend_rate = defend_wl["wins"] / defend_total if defend_total else 0

    if pick_rate >= 0.6 and defend_rate >= 0.6:
        return GREEN_FILL
    if defend_rate <= 0.4 or avoid.get("ratio", 0) >= 0.7:
        return RED_FILL
    return None


def export_map_matrix(conn: sqlite3.Connection, output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Map Matrix"

    teams = _get_all_teams(conn)
    maps = _get_all_maps(conn)

    # Header row
    ws.cell(row=1, column=1, value="Team").fill = HEADER_FILL
    ws.cell(row=1, column=1).font = HEADER_FONT
    for col_idx, (_, map_name, mode) in enumerate(maps, start=2):
        cell = ws.cell(row=1, column=col_idx, value=f"{map_name} ({mode})")
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, (team_id, abbr) in enumerate(teams, start=2):
        ws.cell(row=row_idx, column=1, value=abbr)
        for col_idx, (map_id, _, _) in enumerate(maps, start=2):
            pwl = pick_win_loss(conn, team_id, map_id)
            dwl = defend_win_loss(conn, team_id, map_id)
            avoid = avoidance_index(conn, team_id, map_id)
            tgt = target_index(conn, team_id, map_id)

            text = (
                f"P:{pwl['wins']}-{pwl['losses']} | "
                f"D:{dwl['wins']}-{dwl['losses']} | "
                f"Av:{avoid['ratio']:.0%}(n={avoid['opportunities']}) | "
                f"Tg:{tgt['ratio']:.0%}(n={tgt['opportunities']})"
            )
            cell = ws.cell(row=row_idx, column=col_idx, value=text)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            fill = _cell_color(pwl, dwl, avoid, tgt)
            if fill:
                cell.fill = fill

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    wb.save(output_path)


def _head_to_head(conn: sqlite3.Connection, team_id: int, opp_id: int, map_id: int) -> dict:
    """W-L record between two specific teams on a specific map (all formats)."""
    row = conn.execute(
        """SELECT
               SUM(CASE WHEN mr.winner_team_id = ? THEN 1 ELSE 0 END),
               SUM(CASE WHEN mr.winner_team_id = ? THEN 1 ELSE 0 END)
           FROM map_results mr
           JOIN matches m ON mr.match_id = m.match_id
           WHERE mr.map_id = ?
             AND ((m.team1_id = ? AND m.team2_id = ?) OR (m.team1_id = ? AND m.team2_id = ?))""",
        (team_id, opp_id, map_id, team_id, opp_id, opp_id, team_id),
    ).fetchone()
    return {"wins": row[0] or 0, "losses": row[1] or 0}


def export_matchup_prep(
    conn: sqlite3.Connection, your_team_id: int, opp_team_id: int, output_path: str
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Match-Up Prep"

    maps = _get_all_maps(conn)

    your_abbr = conn.execute("SELECT abbreviation FROM teams WHERE team_id = ?", (your_team_id,)).fetchone()[0]
    opp_abbr = conn.execute("SELECT abbreviation FROM teams WHERE team_id = ?", (opp_team_id,)).fetchone()[0]

    # Header
    headers = ["Map (Mode)", "H2H",
               f"{your_abbr} Pick W-L", f"{your_abbr} Defend W-L",
               f"{opp_abbr} Pick W-L", f"{opp_abbr} Defend W-L",
               f"{opp_abbr} Avoid%", f"{opp_abbr} Target%", "Dominance"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    # Data rows — one per map
    for row_idx, (map_id, map_name, mode) in enumerate(maps, start=2):
        ws.cell(row=row_idx, column=1, value=f"{map_name} ({mode})")

        h2h = _head_to_head(conn, your_team_id, opp_team_id, map_id)
        your_pwl = pick_win_loss(conn, your_team_id, map_id)
        your_dwl = defend_win_loss(conn, your_team_id, map_id)
        opp_pwl = pick_win_loss(conn, opp_team_id, map_id)
        opp_dwl = defend_win_loss(conn, opp_team_id, map_id)
        opp_avoid = avoidance_index(conn, opp_team_id, map_id)
        opp_tgt = target_index(conn, opp_team_id, map_id)
        margins = score_margins(conn, your_team_id, map_id)

        ws.cell(row=row_idx, column=2, value=f"{h2h['wins']}-{h2h['losses']}")
        ws.cell(row=row_idx, column=3, value=f"{your_pwl['wins']}-{your_pwl['losses']}")
        ws.cell(row=row_idx, column=4, value=f"{your_dwl['wins']}-{your_dwl['losses']}")
        ws.cell(row=row_idx, column=5, value=f"{opp_pwl['wins']}-{opp_pwl['losses']}")
        ws.cell(row=row_idx, column=6, value=f"{opp_dwl['wins']}-{opp_dwl['losses']}")
        ws.cell(row=row_idx, column=7, value=f"{opp_avoid['ratio']:.0%} (n={opp_avoid['opportunities']})")
        ws.cell(row=row_idx, column=8, value=f"{opp_tgt['ratio']:.0%} (n={opp_tgt['opportunities']})")

        dom_counts = {}
        for m in margins:
            if m["dominance"]:
                dom_counts[m["dominance"]] = dom_counts.get(m["dominance"], 0) + 1
        dom_str = ", ".join(f"{k}:{v}" for k, v in dom_counts.items()) if dom_counts else "-"
        ws.cell(row=row_idx, column=9, value=dom_str)

        # Yellow fill for low sample
        if opp_avoid["opportunities"] < LOW_SAMPLE_THRESHOLD:
            for c in range(7, 9):
                ws.cell(row=row_idx, column=c).fill = YELLOW_FILL

    # Footer — Elo ratings
    footer_row = len(maps) + 3
    your_elo = get_current_elo(conn, your_team_id)
    opp_elo = get_current_elo(conn, opp_team_id)
    your_lc = " (LOW CONFIDENCE)" if is_low_confidence(conn, your_team_id) else ""
    opp_lc = " (LOW CONFIDENCE)" if is_low_confidence(conn, opp_team_id) else ""

    ws.cell(row=footer_row, column=1, value="Elo Ratings").font = Font(bold=True)
    ws.cell(row=footer_row + 1, column=1, value=f"{your_abbr}: {your_elo:.0f}{your_lc}")
    ws.cell(row=footer_row + 1, column=3, value=f"{opp_abbr}: {opp_elo:.0f}{opp_lc}")

    # Ban summary section
    your_bans = get_ban_summary(conn, your_team_id, opp_team_id)
    opp_bans = get_ban_summary(conn, opp_team_id, your_team_id)

    if your_bans or opp_bans:
        ban_row = footer_row + 3
        ws.cell(row=ban_row, column=1, value="Head-to-Head Ban Data").font = Font(bold=True)
        ban_row += 1

        if your_bans:
            ws.cell(row=ban_row, column=1, value=f"{your_abbr} bans vs {opp_abbr}:")
            ban_strs = [f"{b['map_name']} {b['mode']} ({b['ban_count']}/{b['total_series']})"
                        for b in your_bans]
            ws.cell(row=ban_row, column=2, value=", ".join(ban_strs))
            ban_row += 1

        if opp_bans:
            ws.cell(row=ban_row, column=1, value=f"{opp_abbr} bans vs {your_abbr}:")
            ban_strs = [f"{b['map_name']} {b['mode']} ({b['ban_count']}/{b['total_series']})"
                        for b in opp_bans]
            ws.cell(row=ban_row, column=2, value=", ".join(ban_strs))

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    wb.save(output_path)


def export_team_profile(
    conn: sqlite3.Connection, team_id: int, output_path: str,
    format_filter: str | None = None,
) -> None:
    wb = Workbook()
    ws = wb.active

    abbr = conn.execute("SELECT abbreviation FROM teams WHERE team_id = ?", (team_id,)).fetchone()[0]
    filter_label = format_filter or "All Formats"
    ws.title = f"{abbr} Profile"

    # --- Map W-L Section ---
    headers = ["Map (Mode)", "W", "L", "W%"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    map_wl = get_team_map_wl(conn, team_id, format_filter)
    for row_idx, entry in enumerate(map_wl, start=2):
        w, l = entry["wins"], entry["losses"]
        total = w + l
        ws.cell(row=row_idx, column=1, value=f"{entry['map_name']} ({entry['mode']})")
        ws.cell(row=row_idx, column=2, value=w)
        ws.cell(row=row_idx, column=3, value=l)
        pct = f"{w / total:.0%}" if total else "-"
        ws.cell(row=row_idx, column=4, value=pct)

        if total and w / total >= 0.6:
            for c in range(1, 5):
                ws.cell(row=row_idx, column=c).fill = GREEN_FILL
        elif total and w / total <= 0.4:
            for c in range(1, 5):
                ws.cell(row=row_idx, column=c).fill = RED_FILL

    # --- Totals ---
    total_row = len(map_wl) + 2
    total_w = sum(e["wins"] for e in map_wl)
    total_l = sum(e["losses"] for e in map_wl)
    total_games = total_w + total_l
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=total_w).font = Font(bold=True)
    ws.cell(row=total_row, column=3, value=total_l).font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=f"{total_w / total_games:.0%}" if total_games else "-")

    # --- Format filter label ---
    ws.cell(row=total_row + 1, column=1, value=f"Filter: {filter_label}").font = Font(italic=True)

    # --- Ban Section ---
    ban_data = get_team_ban_summary(conn, team_id)
    if ban_data["total_series"] > 0:
        ban_start = total_row + 3

        ws.cell(row=ban_start, column=1, value=f"{abbr} Bans ({ban_data['total_series']} series)").font = Font(bold=True)
        ban_headers = ["Map (Mode)", "Times Banned", "Ban Rate"]
        ban_start += 1
        for col_idx, h in enumerate(ban_headers, start=1):
            cell = ws.cell(row=ban_start, column=col_idx, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

        for i, b in enumerate(ban_data["team_bans"], start=1):
            r = ban_start + i
            ws.cell(row=r, column=1, value=f"{b['map_name']} ({b['mode']})")
            ws.cell(row=r, column=2, value=b["ban_count"])
            ws.cell(row=r, column=3, value=f"{b['ban_count'] / b['total_series']:.0%}")

        opp_start = ban_start + len(ban_data["team_bans"]) + 2
        ws.cell(row=opp_start, column=1, value=f"Opponent Bans vs {abbr} ({ban_data['total_series']} series)").font = Font(bold=True)
        opp_start += 1
        for col_idx, h in enumerate(ban_headers, start=1):
            cell = ws.cell(row=opp_start, column=col_idx, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

        for i, b in enumerate(ban_data["opponent_bans"], start=1):
            r = opp_start + i
            ws.cell(row=r, column=1, value=f"{b['map_name']} ({b['mode']})")
            ws.cell(row=r, column=2, value=b["ban_count"])
            ws.cell(row=r, column=3, value=f"{b['ban_count'] / b['total_series']:.0%}")

    # --- Elo Footer ---
    elo_row = ws.max_row + 2
    elo = get_current_elo(conn, team_id)
    lc = " (LOW CONFIDENCE)" if is_low_confidence(conn, team_id) else ""
    ws.cell(row=elo_row, column=1, value="Elo Rating").font = Font(bold=True)
    ws.cell(row=elo_row, column=2, value=f"{elo:.0f}{lc}")

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    wb.save(output_path)
